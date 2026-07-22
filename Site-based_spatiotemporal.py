
import pandas as pd
import numpy as np
from xgboost import XGBRegressor
from sklearn.metrics import mean_squared_error
from itertools import product
import os
import shap
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def train_and_predict(train_file_path, test_folder_path, output_folder_path):
    # 读取训练数据
    train_data = pd.read_excel(train_file_path)

    # 假设COMID和date列不需要作为特征
    features = [col for col in train_data.columns if col not in ['COMID', 'date', 'temp']]

    # 检查训练数据中的缺失值
    print("训练数据缺失值统计:")
    print(train_data[features].isnull().sum())

    # 创建缺失值指示特征
    for col in features:
        train_data[f'{col}_missing'] = train_data[col].isnull().astype(int)

    # 直接使用原始数据，保留缺失值
    X = train_data[features]
    y = train_data['temp']

    # 获取COMID的唯一值
    comids = train_data['COMID'].unique()

    # 将COMID均分为6份
    np.random.seed(42)  # 确保可重复性
    np.random.shuffle(comids)
    comid_splits = np.array_split(comids, 6)

    # 将date分为4份
    train_data['date'] = pd.to_datetime(train_data['date'])
    date_splits = [
        (train_data['date'] >= '2019-01-01') & (train_data['date'] <= '2019-06-30'),
        (train_data['date'] >= '2019-07-01') & (train_data['date'] <= '2019-12-31'),
        (train_data['date'] >= '2020-01-01') & (train_data['date'] <= '2020-06-30'),
        (train_data['date'] >= '2020-07-01') & (train_data['date'] <= '2020-12-31')
    ]

    # 定义模型参数的取值范围
    param_grid = {
        'max_depth': [3, 5, 7, 9],
        'learning_rate': [0.01, 0.1, 0.2],
        'n_estimators': [50, 100, 150],
        'missing': [np.nan],
        'tree_method': ['hist'],
        'enable_categorical': [True]
    }

    # 生成所有参数组合
    param_combinations = list(product(*param_grid.values()))

    # 最佳参数和指标初始化
    best_space_val_rmse = float('inf')
    best_time_val_rmse = float('inf')
    best_random_val_rmse = float('inf')

    best_space_params = None
    best_time_params = None
    best_random_params = None

    # ---------------------- 空间交叉验证 ----------------------
    print("\n开始空间交叉验证...")

    # 选择第6份作为测试集，并且在整个参数搜索过程中保持不变
    # test_comids = comid_splits[5]
    # train_val_comids = np.concatenate([comid_splits[i] for i in range(5)])  # 修正：原range(1,6)可能漏第0份，改为range(5)取前5份

    # 选择第5份作为测试集，并且在整个参数搜索过程中保持不变
    # test_comids = comid_splits[4]
    # train_val_comids = np.concatenate([comid_splits[i] for i in [0, 1, 2, 3, 5]])  # 修正：原range(1,6)可能漏第0份，改为range(5)取前5份

    # 选择第4份作为测试集，并且在整个参数搜索过程中保持不变
    test_comids = comid_splits[3]
    train_val_comids = np.concatenate([comid_splits[i] for i in [0, 1, 2, 4, 5]])  # 修正：原range(1,6)可能漏第0份，改为range(5)取前5份

    # # 选择第3份作为测试集，并且在整个参数搜索过程中保持不变
    # test_comids = comid_splits[2]
    # train_val_comids = np.concatenate([comid_splits[i] for i in [0, 1, 3, 4, 5]])  # 修正：原range(1,6)可能漏第0份，改为range(5)取前5份
    #
    # # # 选择第2份作为测试集，并且在整个参数搜索过程中保持不变
    # test_comids = comid_splits[1]
    # train_val_comids = np.concatenate([comid_splits[i] for i in [0, 2, 3, 4, 5]])  # 修正：原range(1,6)可能漏第0份，改为range(5)取前5份
    #
    # # 选择第1份作为测试集，并且在整个参数搜索过程中保持不变
    # test_comids = comid_splits[0]
    # train_val_comids = np.concatenate([comid_splits[i] for i in range(1,6)])  # 修正：原range(1,6)可能漏第0份，改为range(5)取前5份


    test_idx = train_data['COMID'].isin(test_comids)
    X_test_space, y_test_space = X[test_idx], y[test_idx]

    # 剩余5份作为训练和验证数据
    # train_val_comids = np.concatenate([comid_splits[i] for i in range(5)])  # 修正：原range(1,6)可能漏第0份，改为range(5)取前5份
    train_val_idx = train_data['COMID'].isin(train_val_comids)
    X_train_val, y_train_val = X[train_val_idx], y[train_val_idx]




    # 将剩余5份均分为5折
    train_val_splits = np.array_split(train_val_comids, 5)

    # 用于记录所有参数组合的结果（新增：用字典存储每个参数组合的验证集数据）
    param_val_results = {}  # 键：参数组合字符串，值：(val_true_list, val_pred_list)
    all_results = []
    best_space_val_rmse = float('inf')
    best_space_params = None

    for params in param_combinations:
        param_dict = dict(zip(param_grid.keys(), params))
        param_str = str(param_dict)  # 用字符串作为键，确保唯一性
        val_rmse_list = []  # 验证集RMSE列表
        train_rmse_list = []  # 训练集RMSE列表
        # 初始化当前参数组合的验证集数据存储
        param_val_results[param_str] = ([], [])  # (真实值列表, 预测值列表)

        # 5折交叉验证
        for val_fold in range(5):
            # 划分验证集
            val_comids = train_val_splits[val_fold]
            val_idx = train_data['COMID'].isin(val_comids)

            # 划分训练集
            train_comids = np.concatenate([train_val_splits[i] for i in range(5) if i != val_fold])
            train_idx = train_data['COMID'].isin(train_comids)

            # 获取训练集、验证集数据
            X_train, y_train = X[train_idx], y[train_idx]
            X_val, y_val = X[val_idx], y[val_idx]

            # 训练模型
            model = XGBRegressor(**param_dict)
            model.fit(X_train, y_train)

            # 在验证集上评估
            y_val_pred = model.predict(X_val)
            val_rmse = np.sqrt(mean_squared_error(y_val, y_val_pred))
            val_rmse_list.append(val_rmse)

            pd.DataFrame({
                '真实值': y_val,
                '预测值': y_val_pred
            }).to_csv('rmse_check.csv', index=False)

            # 在训练集上评估
            y_train_pred = model.predict(X_train)
            train_rmse = np.sqrt(mean_squared_error(y_train, y_train_pred))
            train_rmse_list.append(train_rmse)

            # 记录当前折的结果
            fold_result = {
                '参数组合': param_str,
                '折数': val_fold + 1,
                '训练集RMSE': train_rmse,
                '验证集RMSE': val_rmse
            }
            all_results.append(fold_result)

            # 保存当前参数组合的验证集真实值和预测值（关键修改：不再依赖best_space_params，直接存入对应参数的列表）
            val_true_list, val_pred_list = param_val_results[param_str]
            val_true_list.extend(y_val.tolist())
            val_pred_list.extend(y_val_pred.tolist())

        # 计算当前参数组合的平均RMSE
        avg_val_rmse = np.mean(val_rmse_list)
        avg_train_rmse = np.mean(train_rmse_list)

        # 记录当前参数组合的平均结果
        param_avg_result = {
            '参数组合': param_str,
            '平均训练集RMSE': avg_train_rmse,
            '平均验证集RMSE': avg_val_rmse
        }
        all_results.append(param_avg_result)

        # 更新最佳参数
        if avg_val_rmse < best_space_val_rmse:
            best_space_val_rmse = avg_val_rmse
            best_space_params = param_dict

    # 提取最优参数对应的验证集数据（关键修复：从字典中获取，避免空列表）
    best_param_str = str(best_space_params)
    if best_param_str in param_val_results:
        best_val_true, best_val_pred = param_val_results[best_param_str]
        # 确保有数据（避免极端情况）
        if len(best_val_true) == 0:
            best_val_rmse = np.nan
            print("警告：最优参数未找到对应的验证集数据，可能是参数组合匹配失败")
        else:
            best_val_rmse = np.sqrt(mean_squared_error(best_val_true, best_val_pred))
    else:
        best_val_rmse = np.nan
        print("警告：未找到最优参数对应的验证集数据")

    # 使用最佳参数在所有训练数据上训练最终模型
    best_space_model = XGBRegressor(**best_space_params)
    best_space_model.fit(X_train_val, y_train_val)

    # 在测试集上评估
    y_test_pred_space = best_space_model.predict(X_test_space)
    best_space_test_rmse = np.sqrt(mean_squared_error(y_test_space, y_test_pred_space))

    # 在训练集上评估
    y_train_pred_space = best_space_model.predict(X_train_val)
    best_space_train_rmse = np.sqrt(mean_squared_error(y_train_val, y_train_pred_space))

    # 输出到Excel
    results_df = pd.DataFrame(all_results)

    # 创建最终结果DataFrame
    final_results = pd.DataFrame({
        '评估数据集': ['训练集', '验证集', '测试集'],
        'RMSE': [best_space_train_rmse, best_val_rmse, best_space_test_rmse]
    })

    # 确保输出文件夹存在
    # excel_output_path = os.path.join(output_folder_path, '模型评估结果-0.xlsx')
    # os.makedirs(os.path.dirname(excel_output_path), exist_ok=True)
    #
    # # 将结果写入Excel
    # with pd.ExcelWriter(excel_output_path, engine='openpyxl') as writer:
    #     results_df.to_excel(writer, sheet_name='所有参数结果', index=False)
    #     final_results.to_excel(writer, sheet_name='最优参数最终评估', index=False)
    #     pd.DataFrame([best_space_params]).to_excel(writer, sheet_name='最优参数', index=False)
    #
    # print(f"结果已保存至: {excel_output_path}")
    print(f"最优参数: {best_space_params}")
    print(f"训练集RMSE: {best_space_train_rmse:.4f}")
    print(f"验证集RMSE: {best_val_rmse:.4f}")
    print(f"测试集RMSE: {best_space_test_rmse:.4f}")

    # ---------------------- 时间交叉验证 ----------------------
    print("\n开始时间交叉验证...")

    # 选择最后一份作为测试集
    test_dates = date_splits[3]
    test_idx = test_dates
    X_test_time, y_test_time = X[test_idx], y[test_idx]

    # 剩余3份作为训练和验证数据
    train_val_dates = [date_splits[i] for i in range(3)]  # 取前3份

    for params in param_combinations:
        param_dict = dict(zip(param_grid.keys(), params))
        val_rmse_list = []  # 验证集RMSE列表

        # 3折交叉验证
        for val_fold in range(3):
            # 划分验证集
            val_dates = train_val_dates[val_fold]
            val_idx = val_dates

            # 划分训练集
            train_dates = [train_val_dates[i] for i in range(3) if i != val_fold]
            train_idx = train_dates[0] | train_dates[1]

            # 获取训练集、验证集数据
            X_train, y_train = X[train_idx], y[train_idx]
            X_val, y_val = X[val_idx], y[val_idx]

            # 训练模型
            model = XGBRegressor(**param_dict)
            model.fit(X_train, y_train)

            # 在验证集上评估
            y_val_pred = model.predict(X_val)
            val_rmse = np.sqrt(mean_squared_error(y_val, y_val_pred))
            val_rmse_list.append(val_rmse)

        # 计算当前参数组合的平均验证集RMSE
        avg_val_rmse = np.mean(val_rmse_list)

        # 更新最佳参数
        if avg_val_rmse < best_time_val_rmse:
            best_time_val_rmse = avg_val_rmse
            best_time_params = param_dict

    # 使用最佳参数在所有训练数据上训练最终模型
    train_val_idx = train_val_dates[0] | train_val_dates[1] | train_val_dates[2]
    best_time_model = XGBRegressor(**best_time_params)
    best_time_model.fit(X[train_val_idx], y[train_val_idx])

    # 在测试集上评估
    y_test_pred_time = best_time_model.predict(X_test_time)
    best_time_test_rmse = np.sqrt(mean_squared_error(y_test_time, y_test_pred_time))

    # 在训练集上评估
    y_train_pred_time = best_time_model.predict(X[train_val_idx])
    best_time_train_rmse = np.sqrt(mean_squared_error(y[train_val_idx], y_train_pred_time))

    # ---------------------- 随机交叉验证 ----------------------
    print("\n开始随机交叉验证...")

    # 随机选择20%数据作为测试集
    np.random.seed(42)
    test_idx = np.random.choice(len(X), int(0.2 * len(X)), replace=False)
    test_mask = np.isin(np.arange(len(X)), test_idx)
    X_test_random, y_test_random = X[test_mask], y[test_mask]

    # 剩余80%作为训练和验证数据
    train_val_mask = ~test_mask
    X_train_val_random, y_train_val_random = X[train_val_mask], y[train_val_mask]

    for params in param_combinations:
        param_dict = dict(zip(param_grid.keys(), params))
        val_rmse_list = []  # 验证集RMSE列表

        # 5折交叉验证
        for _ in range(5):
            # 划分验证集（20% of 训练集）
            val_idx = np.random.choice(len(X_train_val_random), int(0.2 * len(X_train_val_random)), replace=False)
            val_mask = np.isin(np.arange(len(X_train_val_random)), val_idx)

            # 获取验证集和训练集
            X_val, y_val = X_train_val_random[val_mask], y_train_val_random[val_mask]
            train_mask = ~val_mask
            X_train, y_train = X_train_val_random[train_mask], y_train_val_random[train_mask]

            # 训练模型
            model = XGBRegressor(**param_dict)
            model.fit(X_train, y_train)

            # 在验证集上评估
            y_val_pred = model.predict(X_val)
            val_rmse = np.sqrt(mean_squared_error(y_val, y_val_pred))
            val_rmse_list.append(val_rmse)

        # 计算平均验证集RMSE
        avg_val_rmse = np.mean(val_rmse_list)

        # 更新最佳参数
        if avg_val_rmse < best_random_val_rmse:
            best_random_val_rmse = avg_val_rmse
            best_random_params = param_dict

    # 使用最佳参数训练最终模型
    best_random_model = XGBRegressor(**best_random_params)
    best_random_model.fit(X_train_val_random, y_train_val_random)

    # 在测试集上评估
    y_test_pred_random = best_random_model.predict(X_test_random)
    best_random_test_rmse = np.sqrt(mean_squared_error(y_test_random, y_test_pred_random))

    # 在训练集上评估
    y_train_pred_random = best_random_model.predict(X_train_val_random)
    best_random_train_rmse = np.sqrt(mean_squared_error(y_train_val_random, y_train_pred_random))

    # 确保结果文件夹存在
    os.makedirs(output_folder_path, exist_ok=True)

    # 读取测试数据并预测
    test_files = [f for f in os.listdir(test_folder_path) if f.endswith('.xlsx')]

    for file in test_files:
        try:
            test_data = pd.read_excel(os.path.join(test_folder_path, file))

            # 检查测试数据中的缺失值
            print(f"\n测试文件 {file} 缺失值统计:")
            missing_counts = test_data[features].isnull().sum()
            print(missing_counts[missing_counts > 0])

            # 创建测试数据的缺失值指示特征
            for col in features:
                if col in test_data.columns:
                    test_data[f'{col}_missing'] = test_data[col].isnull().astype(int)
                else:
                    test_data[f'{col}_missing'] = 1  # 特征不存在时标记为缺失

            # 确保测试特征与训练一致
            X_test = test_data[features]

            # 使用最佳模型进行预测
            predicted_rwt_space = best_space_model.predict(X_test)
            predicted_rwt_time = best_time_model.predict(X_test)
            predicted_rwt_random = best_random_model.predict(X_test)

            # 保存预测结果
            test_data['Predicted_RWT_Space'] = predicted_rwt_space
            test_data['Predicted_RWT_Time'] = predicted_rwt_time
            test_data['Predicted_RWT_Random'] = predicted_rwt_random
            output_path = os.path.join(output_folder_path, file)
            test_data.to_excel(output_path, index=False)
            print(f"已处理并保存: {output_path}")

        except Exception as e:
            print(f"处理文件 {file} 时出错: {str(e)}")

    # 输出所有交叉验证结果
    print("\n" + "=" * 50)
    print("三种交叉验证方式下的三数据集RMSE结果")
    print("=" * 50)

    print("\n【空间交叉验证】")
    print(f"训练集RMSE: {best_space_train_rmse:.4f}")
    print(f"验证集RMSE: {best_val_rmse:.4f}")
    print(f"测试集RMSE: {best_space_test_rmse:.4f}")
    print(f"最佳参数: {best_space_params}")

    print("\n【时间交叉验证】")
    print(f"训练集RMSE: {best_time_train_rmse:.4f}")
    print(f"验证集RMSE: {best_time_val_rmse:.4f}")
    print(f"测试集RMSE: {best_time_test_rmse:.4f}")
    print(f"最佳参数: {best_time_params}")

    print("\n【随机交叉验证】")
    print(f"训练集RMSE: {best_random_train_rmse:.4f}")
    print(f"验证集RMSE: {best_random_val_rmse:.4f}")
    print(f"测试集RMSE: {best_random_test_rmse:.4f}")
    print(f"最佳参数: {best_random_params}")

    # 计算特征贡献度
    calculate_feature_contributions(
        X, best_space_model, best_time_model, best_random_model, output_folder_path
    )


def calculate_feature_contributions(X, space_model, time_model, random_model, output_folder):
    print("\n正在计算特征贡献度...")
    try:
        # 初始化SHAP解释器
        space_explainer = shap.TreeExplainer(space_model)
        time_explainer = shap.TreeExplainer(time_model)
        random_explainer = shap.TreeExplainer(random_model)

        # 计算SHAP值（使用前1000样本加速计算，大规模数据可抽样）
        sample_idx = np.random.choice(len(X), min(1000, len(X)), replace=False)
        X_sample = X.iloc[sample_idx] if isinstance(X, pd.DataFrame) else X[sample_idx]

        space_shap_values = space_explainer.shap_values(X_sample)
        time_shap_values = time_explainer.shap_values(X_sample)
        random_shap_values = random_explainer.shap_values(X_sample)

        # 整理特征贡献度结果
        feature_contributions = pd.DataFrame({
            'Feature': X.columns,
            'Space_Mean_Abs_SHAP': np.abs(space_shap_values).mean(axis=0),
            'Space_Rank': np.argsort(np.argsort(-np.abs(space_shap_values).mean(axis=0))) + 1,
            'Space_Direction': np.sign(space_shap_values).mean(axis=0),
            'Time_Mean_Abs_SHAP': np.abs(time_shap_values).mean(axis=0),
            'Time_Rank': np.argsort(np.argsort(-np.abs(time_shap_values).mean(axis=0))) + 1,
            'Time_Direction': np.sign(time_shap_values).mean(axis=0),
            'Random_Mean_Abs_SHAP': np.abs(random_shap_values).mean(axis=0),
            'Random_Rank': np.argsort(np.argsort(-np.abs(random_shap_values).mean(axis=0))) + 1,
            'Random_Direction': np.sign(random_shap_values).mean(axis=0),
        })

        # 计算综合贡献度
        feature_contributions['Mean_Contribution'] = (
                                                             feature_contributions['Space_Mean_Abs_SHAP'] +
                                                             feature_contributions['Time_Mean_Abs_SHAP'] +
                                                             feature_contributions['Random_Mean_Abs_SHAP']
                                                     ) / 3
        feature_contributions['Overall_Rank'] = (
                np.argsort(np.argsort(-feature_contributions['Mean_Contribution'])) + 1
        )
        total_contribution = feature_contributions['Mean_Contribution'].sum()
        feature_contributions['Contribution_Percentage'] = (
                feature_contributions['Mean_Contribution'] / total_contribution * 100
        )

        # 保存特征贡献度结果
        output_path = os.path.join(output_folder, 'Feature_Contributions.xlsx')
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            feature_contributions.to_excel(writer, sheet_name='Contributions', index=False)
            workbook = writer.book
            sheet = workbook['Contributions']
            # 美化表格
            header_font = Font(bold=True)
            for cell in sheet[1]:
                cell.font = header_font
            for col in sheet.columns:
                max_len = max(len(str(cell.value)) for cell in col) + 2
                sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max_len, 30)

        print(f"特征贡献度已保存到: {output_path}")
        generate_contribution_plots(feature_contributions, output_folder)

    except Exception as e:
        print(f"计算特征贡献度时出错: {str(e)}")


def generate_contribution_plots(contributions, output_folder):
    plot_folder = os.path.join(output_folder, 'Contribution_Plots')
    os.makedirs(plot_folder, exist_ok=True)
    top_n = min(20, len(contributions))
    top_features = contributions.sort_values('Mean_Contribution', ascending=False).head(top_n)

    # 总体贡献度图
    plt.figure(figsize=(10, 8))
    colors = np.where(top_features['Space_Direction'] > 0, '#3498db', '#e74c3c')
    plt.barh(top_features['Feature'], top_features['Mean_Contribution'], color=colors)
    plt.xlabel('平均SHAP值绝对值（贡献度）')
    plt.ylabel('特征')
    plt.title('特征总体贡献度排名（Top 20）')
    plt.grid(axis='x', linestyle='--', alpha=0.7)
    plt.tight_layout()
    plt.savefig(os.path.join(plot_folder, 'Overall_Contribution.png'), dpi=300)
    plt.close()

    # 分模型贡献度图
    for model_type in ['Space', 'Time', 'Random']:
        plt.figure(figsize=(10, 8))
        colors = np.where(top_features[f'{model_type}_Direction'] > 0, '#3498db', '#e74c3c')
        plt.barh(
            top_features['Feature'],
            top_features[f'{model_type}_Mean_Abs_SHAP'],
            color=colors
        )
        plt.xlabel('平均SHAP值绝对值（贡献度）')
        plt.ylabel('特征')
        plt.title(f'{model_type}模型特征贡献度排名（Top 20）')
        plt.grid(axis='x', linestyle='--', alpha=0.7)
        plt.tight_layout()
        plt.savefig(os.path.join(plot_folder, f'{model_type}_Contribution.png'), dpi=600)
        plt.close()


if __name__ == "__main__":
    # 路径配置
    train_file_path = "E:\\huai_river\\Station_RWT\\Extracted_matching_with_temp.xlsx"
    test_folder_path = 'E:\\huai_river\\Huairiver_GEE_data\\Daily_data\\old\\11-test-1'
    output_folder_path = "E:\huai_river\Huairiver_GEE_data\Daily_data\old\\result-0718"
    # train_and_predict(train_file_path, test_folder_path, output_folder_path)



if __name__ == "__main__":
    # 路径配置
    train_file_path = "E:\\huai_river\\Station_RWT\\Extracted_matching_with_temp_LST.xlsx"
    test_folder_path = 'E:\\Global_river_temp\\river_extent\\Five_zone\\binary_LST_value\\Order2_big'
    output_folder_path = "E:\Global_river_temp\\river_extent\\Five_zone\\Result\\Order2_big"
    # train_and_predict(train_file_path, test_folder_path, output_folder_path)



# if __name__ == "__main__":
#     # 路径配置
#     train_file_path = "E:\\huai_river\\Station_RWT\\Extracted_matching_with_temp_LST.xlsx"
#     test_folder_path = 'E:\\Global_river_temp\\river_extent\\Five_zone\\binary_LST_value\\Order3_big'
#     output_folder_path = "E:\Global_river_temp\\river_extent\\Five_zone\\Result\\Order3_big"
#     train_and_predict(train_file_path, test_folder_path, output_folder_path)
#
#
#
# if __name__ == "__main__":
#     # 路径配置
#     train_file_path = "E:\\huai_river\\Station_RWT\\Extracted_matching_with_temp_LST.xlsx"
#     test_folder_path = 'E:\\Global_river_temp\\river_extent\\Five_zone\\binary_LST_value\\Order4_big'
#     output_folder_path = "E:\Global_river_temp\\river_extent\\Five_zone\\Result\\Order4_big"
#     train_and_predict(train_file_path, test_folder_path, output_folder_path)
#
#
#
# if __name__ == "__main__":
#     # 路径配置
#     train_file_path = "E:\\huai_river\\Station_RWT\\Extracted_matching_with_temp_LST.xlsx"
#     test_folder_path = 'E:\\Global_river_temp\\river_extent\\Five_zone\\binary_LST_value\\Order5_big'
#     output_folder_path = "E:\Global_river_temp\\river_extent\\Five_zone\\Result\\Order5_big"
#     train_and_predict(train_file_path, test_folder_path, output_folder_path)
#
#
#
# if __name__ == "__main__":
#     # 路径配置
#     train_file_path = "E:\\huai_river\\Station_RWT\\Extracted_matching_with_temp_LST.xlsx"
#     test_folder_path = 'E:\\Global_river_temp\\river_extent\\Five_zone\\binary_LST_value\\Order6_big'
#     output_folder_path = "E:\Global_river_temp\\river_extent\\Five_zone\\Result\\Order6_big"
#     train_and_predict(train_file_path, test_folder_path, output_folder_path)








# *************************************  导出 测试、独立验证的样本数据   **************************************************
# ========================================    在原始xlsx内标注  =======================================================
# =========================================  包含COMID 唯一值  =================================================

# =============================   输出交叉验证过程中的每一折验证集的预测值 样本集  ============================================
import pandas as pd
import numpy as np
from xgboost import XGBRegressor
from sklearn.metrics import mean_squared_error
from itertools import product
import os
import shap
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def train_and_predict(train_file_path, test_folder_path, output_folder_path):
    # 目标参数组合（用户指定）
    target_params = {
        # 'max_depth': 7,
        # 'learning_rate': 0.1,
        # 'n_estimators': 150,
        # 'missing': np.nan,
        # 'tree_method': 'hist',
        # 'enable_categorical': True

        'max_depth': [3, 5, 7, 9],
        'learning_rate': [0.01, 0.1, 0.2],
        'n_estimators': [150, 200, 300, 400],
        'min_child_weight':[3,5,7,10],
        'missing': [np.nan],
        'tree_method': ['hist'],
        'enable_categorical': [True]
    }



    # 读取训练数据（保留原始数据用于导出和标注）
    train_data_original = pd.read_excel(train_file_path)
    # 复制一份用于特征处理（不修改原始数据）
    train_data = train_data_original.copy()

    # 假设COMID和date列不需要作为特征
    features = [col for col in train_data.columns if col not in ['COMID', 'date', 'temp']]

    # 检查训练数据中的缺失值
    print("训练数据缺失值统计:")
    print(train_data[features].isnull().sum())

    # 创建缺失值指示特征（仅用于模型训练，不影响原始数据导出）
    for col in features:
        train_data[f'{col}_missing'] = train_data[col].isnull().astype(int)

    # 直接使用原始数据，保留缺失值
    X = train_data[features]
    y = train_data['temp']

    # 获取COMID的唯一值
    comids = train_data['COMID'].unique()

    # 将COMID均分为6份
    np.random.seed(42)  # 确保可重复性
    np.random.shuffle(comids)
    comid_splits = np.array_split(comids, 6)

    # 将date分为4份
    train_data['date'] = pd.to_datetime(train_data['date'])
    date_splits = [
        (train_data['date'] >= '2019-01-01') & (train_data['date'] <= '2019-06-30'),
        (train_data['date'] >= '2019-07-01') & (train_data['date'] <= '2019-12-31'),
        (train_data['date'] >= '2020-01-01') & (train_data['date'] <= '2020-06-30'),
        (train_data['date'] >= '2020-07-01') & (train_data['date'] <= '2020-12-31')
    ]

    # 定义模型参数的取值范围
    param_grid = {
        'max_depth': [3, 5, 7, 9],
        'learning_rate': [0.01, 0.1, 0.2],
        'n_estimators': [150, 200, 300, 400],
        'min_child_weight':[3,5,7,10],
        'missing': [np.nan],
        'tree_method': ['hist'],
        'enable_categorical': [True]
    }

    # 生成所有参数组合
    param_combinations = list(product(*param_grid.values()))

    # 最佳参数和指标初始化
    best_space_val_rmse = float('inf')
    best_time_val_rmse = float('inf')
    best_random_val_rmse = float('inf')

    best_space_params = None
    best_time_params = None
    best_random_params = None

    # ---------------------- 空间交叉验证 ----------------------
    print("\n开始空间交叉验证...")

    # 选择第4份作为测试集，并且在整个参数搜索过程中保持不变
    test_comids = comid_splits[3]
    train_val_comids = np.concatenate([comid_splits[i] for i in [0, 1, 2, 4, 5]])

    # 划分测试集和训练验证集索引（基于原始数据）
    test_idx = train_data_original['COMID'].isin(test_comids)
    train_val_idx = train_data_original['COMID'].isin(train_val_comids)

    # 创建标注列（不在原始数据上直接修改，而是创建副本）
    labeled_data = train_data_original.copy()

    # 添加测试集标注列
    labeled_data['is_test'] = 0
    labeled_data.loc[test_idx, 'is_test'] = 1

    # 添加各折验证集标注列
    for i in range(5):
        labeled_data[f'is_validation{i + 1}'] = 0

    # 导出空间交叉验证的测试集样本（使用原始数据列）
    test_samples = train_data_original[test_idx].copy()
    test_export_path = os.path.join(output_folder_path, '空间交叉验证_测试集样本.xlsx')
    test_samples.to_excel(test_export_path, index=False)
    print(f"已导出空间交叉验证测试集样本至: {test_export_path}")

    # 将剩余5份均分为5折
    train_val_splits = np.array_split(train_val_comids, 5)

    # 导出每折验证集样本并标记
    for val_fold in range(5):
        val_comids = train_val_splits[val_fold]
        val_idx_original = train_data_original['COMID'].isin(val_comids)

        # 标记当前折的验证集
        labeled_data.loc[val_idx_original, f'is_validation{val_fold + 1}'] = 1

        val_samples = train_data_original[val_idx_original].copy()
        val_export_path = os.path.join(output_folder_path, f'空间交叉验证_验证集样本_第{val_fold + 1}折.xlsx')
        val_samples.to_excel(val_export_path, index=False)
        print(f"已导出空间交叉验证第{val_fold + 1}折验证集样本至: {val_export_path}")

    # 保存带标注的原始数据副本
    labeled_export_path = os.path.join(output_folder_path, 'Extracted_matching_with_temp_with_labels.xlsx')
    labeled_data.to_excel(labeled_export_path, index=False)
    print(f"已保存带标注的原始数据至: {labeled_export_path}")

    # 收集各数据集的COMID信息并导出
    comid_info = []

    # 添加测试集COMID
    for comid in test_comids:
        comid_info.append({
            'COMID': comid,
            '数据集类型': 'test',
            '数据集编号': 3,  # 3表示测试集
            '折数': 'N/A'
        })

    # 添加验证集COMID
    for fold, val_comids in enumerate(train_val_splits, 1):
        for comid in val_comids:
            comid_info.append({
                'COMID': comid,
                '数据集类型': 'validation',
                '数据集编号': 2,  # 2表示验证集
                '折数': fold
            })

    # 添加训练集COMID（所有不在测试集和当前折验证集的COMID）
    for fold, val_comids in enumerate(train_val_splits, 1):
        # 训练集是除了测试集和当前折验证集之外的COMID
        train_comids = np.concatenate([train_val_splits[i] for i in range(5) if i != fold - 1])
        for comid in train_comids:
            comid_info.append({
                'COMID': comid,
                '数据集类型': 'train',
                '数据集编号': 1,  # 1表示训练集
                '折数': fold
            })

    # 创建COMID分配数据框并去重（每个折的分配单独记录）
    comid_df = pd.DataFrame(comid_info)

    # 导出COMID分配信息
    comid_export_path = os.path.join(output_folder_path, '空间交叉验证_COMID分配.xlsx')
    with pd.ExcelWriter(comid_export_path, engine='openpyxl') as writer:
        comid_df.to_excel(writer, sheet_name='COMID分配', index=False)

        # 美化表格
        workbook = writer.book
        sheet = workbook['COMID分配']
        header_font = Font(bold=True)
        for cell in sheet[1]:
            cell.font = header_font
        for col in sheet.columns:
            max_len = max(len(str(cell.value)) for cell in col) + 2
            sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max_len, 30)

    print(f"已导出空间交叉验证COMID分配信息至: {comid_export_path}")

    # 用于模型训练的测试集和训练验证集
    X_test_space, y_test_space = X[test_idx], y[test_idx]
    X_train_val, y_train_val = X[train_val_idx], y[train_val_idx]

    # 用于记录所有参数组合的结果
    param_val_results = {}  # 键：参数组合字符串，值：(val_true_list, val_pred_list)
    all_results = []
    best_space_val_rmse = float('inf')
    best_space_params = None

    # 创建目标参数组合的输出文件夹
    target_params_folder = os.path.join(output_folder_path, "目标参数组合结果")
    os.makedirs(target_params_folder, exist_ok=True)

    for params in param_combinations:
        param_dict = dict(zip(param_grid.keys(), params))
        param_str = str(param_dict)  # 用字符串作为键，确保唯一性
        val_rmse_list = []  # 验证集RMSE列表
        train_rmse_list = []  # 训练集RMSE列表
        param_val_results[param_str] = ([], [])  # (真实值列表, 预测值列表)

        # 5折交叉验证
        for val_fold in range(5):
            # 划分验证集
            val_comids = train_val_splits[val_fold]
            val_idx = train_data['COMID'].isin(val_comids)

            # 划分训练集
            train_comids = np.concatenate([train_val_splits[i] for i in range(5) if i != val_fold])
            train_idx = train_data['COMID'].isin(train_comids)

            # 获取训练集、验证集数据
            X_train, y_train = X[train_idx], y[train_idx]
            X_val, y_val = X[val_idx], y[val_idx]

            # 训练模型
            model = XGBRegressor(**param_dict)
            model.fit(X_train, y_train)

            # 在验证集上评估
            y_val_pred = model.predict(X_val)
            val_rmse = np.sqrt(mean_squared_error(y_val, y_val_pred))
            val_rmse_list.append(val_rmse)

            # 检查是否为目标参数组合，如果是则保存该折的验证集样本及模拟值
            if param_dict == target_params:
                # 获取该折验证集的原始数据
                val_original_data = train_data_original[train_data_original['COMID'].isin(val_comids)].copy()
                # 确保索引对齐（原始数据和特征数据的索引可能不同）
                val_original_data = val_original_data.reset_index(drop=True)
                # 获取对应预测值（按原始数据顺序）
                val_idx_original = train_data_original['COMID'].isin(val_comids)
                val_features = X[val_idx_original].copy().reset_index(drop=True)
                val_pred = model.predict(val_features)
                # 添加模拟值列
                val_original_data['模拟值'] = val_pred
                # 保存结果
                target_val_path = os.path.join(target_params_folder,
                                               f'目标参数组合_第{val_fold + 1}折验证集及模拟值.xlsx')
                val_original_data.to_excel(target_val_path, index=False)
                print(f"已保存目标参数组合第{val_fold + 1}折验证集及模拟值至: {target_val_path}")

            # 在训练集上评估
            y_train_pred = model.predict(X_train)
            train_rmse = np.sqrt(mean_squared_error(y_train, y_train_pred))
            train_rmse_list.append(train_rmse)

            # 记录当前折的结果
            fold_result = {
                '参数组合': param_str,
                '折数': val_fold + 1,
                '训练集RMSE': train_rmse,
                '验证集RMSE': val_rmse
            }
            all_results.append(fold_result)

            # 保存当前参数组合的验证集真实值和预测值
            val_true_list, val_pred_list = param_val_results[param_str]
            val_true_list.extend(y_val.tolist())
            val_pred_list.extend(y_val_pred.tolist())

        # 计算当前参数组合的平均RMSE
        avg_val_rmse = np.mean(val_rmse_list)
        avg_train_rmse = np.mean(train_rmse_list)

        # 记录当前参数组合的平均结果
        param_avg_result = {
            '参数组合': param_str,
            '平均训练集RMSE': avg_train_rmse,
            '平均验证集RMSE': avg_val_rmse
        }
        all_results.append(param_avg_result)

        # 更新最佳参数
        if avg_val_rmse < best_space_val_rmse:
            best_space_val_rmse = avg_val_rmse
            best_space_params = param_dict

    # 提取最优参数对应的验证集数据
    best_param_str = str(best_space_params)
    if best_param_str in param_val_results:
        best_val_true, best_val_pred = param_val_results[best_param_str]
        if len(best_val_true) == 0:
            best_val_rmse = np.nan
            print("警告：最优参数未找到对应的验证集数据，可能是参数组合匹配失败")
        else:
            best_val_rmse = np.sqrt(mean_squared_error(best_val_true, best_val_pred))
    else:
        best_val_rmse = np.nan
        print("警告：未找到最优参数对应的验证集数据")

    # 使用最佳参数在所有训练数据上训练最终模型
    best_space_model = XGBRegressor(**best_space_params)
    best_space_model.fit(X_train_val, y_train_val)

    # 在测试集上评估
    y_test_pred_space = best_space_model.predict(X_test_space)
    best_space_test_rmse = np.sqrt(mean_squared_error(y_test_space, y_test_pred_space))

    # 在训练集上评估
    y_train_pred_space = best_space_model.predict(X_train_val)
    best_space_train_rmse = np.sqrt(mean_squared_error(y_train_val, y_train_pred_space))

    # 输出到Excel
    results_df = pd.DataFrame(all_results)

    # 创建最终结果DataFrame
    final_results = pd.DataFrame({
        '评估数据集': ['训练集', '验证集', '测试集'],
        'RMSE': [best_space_train_rmse, best_val_rmse, best_space_test_rmse]
    })

    # 确保输出文件夹存在
    excel_output_path = os.path.join(output_folder_path, '模型评估结果-0.xlsx')
    os.makedirs(os.path.dirname(excel_output_path), exist_ok=True)

    # 将结果写入Excel
    with pd.ExcelWriter(excel_output_path, engine='openpyxl') as writer:
        results_df.to_excel(writer, sheet_name='所有参数结果', index=False)
        final_results.to_excel(writer, sheet_name='最优参数最终评估', index=False)
        pd.DataFrame([best_space_params]).to_excel(writer, sheet_name='最优参数', index=False)

    print(f"结果已保存至: {excel_output_path}")
    print(f"最优参数: {best_space_params}")
    print(f"训练集RMSE: {best_space_train_rmse:.4f}")
    print(f"验证集RMSE: {best_val_rmse:.4f}")
    print(f"测试集RMSE: {best_space_test_rmse:.4f}")

    # ---------------------- 时间交叉验证 ----------------------
    print("\n开始时间交叉验证...")

    # 选择最后一份作为测试集
    test_dates = date_splits[3]
    test_idx = test_dates
    X_test_time, y_test_time = X[test_idx], y[test_idx]

    # 剩余3份作为训练和验证数据
    train_val_dates = [date_splits[i] for i in range(3)]  # 取前3份

    for params in param_combinations:
        param_dict = dict(zip(param_grid.keys(), params))
        val_rmse_list = []  # 验证集RMSE列表

        # 3折交叉验证
        for val_fold in range(3):
            # 划分验证集
            val_dates = train_val_dates[val_fold]
            val_idx = val_dates

            # 划分训练集
            train_dates = [train_val_dates[i] for i in range(3) if i != val_fold]
            train_idx = train_dates[0] | train_dates[1]

            # 获取训练集、验证集数据
            X_train, y_train = X[train_idx], y[train_idx]
            X_val, y_val = X[val_idx], y[val_idx]

            # 训练模型
            model = XGBRegressor(**param_dict)
            model.fit(X_train, y_train)

            # 在验证集上评估
            y_val_pred = model.predict(X_val)
            val_rmse = np.sqrt(mean_squared_error(y_val, y_val_pred))
            val_rmse_list.append(val_rmse)

        # 计算当前参数组合的平均验证集RMSE
        avg_val_rmse = np.mean(val_rmse_list)

        # 更新最佳参数
        if avg_val_rmse < best_time_val_rmse:
            best_time_val_rmse = avg_val_rmse
            best_time_params = param_dict

    # 使用最佳参数在所有训练数据上训练最终模型
    train_val_idx = train_val_dates[0] | train_val_dates[1] | train_val_dates[2]
    best_time_model = XGBRegressor(**best_time_params)
    best_time_model.fit(X[train_val_idx], y[train_val_idx])

    # 在测试集上评估
    y_test_pred_time = best_time_model.predict(X_test_time)
    best_time_test_rmse = np.sqrt(mean_squared_error(y_test_time, y_test_pred_time))

    # 在训练集上评估
    y_train_pred_time = best_time_model.predict(X[train_val_idx])
    best_time_train_rmse = np.sqrt(mean_squared_error(y[train_val_idx], y_train_pred_time))

    # ---------------------- 随机交叉验证 ----------------------
    print("\n开始随机交叉验证...")

    # 随机选择20%数据作为测试集
    np.random.seed(42)
    test_idx = np.random.choice(len(X), int(0.2 * len(X)), replace=False)
    test_mask = np.isin(np.arange(len(X)), test_idx)
    X_test_random, y_test_random = X[test_mask], y[test_mask]

    # 剩余80%作为训练和验证数据
    train_val_mask = ~test_mask
    X_train_val_random, y_train_val_random = X[train_val_mask], y[train_val_mask]

    for params in param_combinations:
        param_dict = dict(zip(param_grid.keys(), params))
        val_rmse_list = []  # 验证集RMSE列表

        # 5折交叉验证
        for _ in range(5):
            # 划分验证集（20% of 训练集）
            val_idx = np.random.choice(len(X_train_val_random), int(0.2 * len(X_train_val_random)), replace=False)
            val_mask = np.isin(np.arange(len(X_train_val_random)), val_idx)

            # 获取验证集和训练集
            X_val, y_val = X_train_val_random[val_mask], y_train_val_random[val_mask]
            train_mask = ~val_mask
            X_train, y_train = X_train_val_random[train_mask], y_train_val_random[train_mask]

            # 训练模型
            model = XGBRegressor(**param_dict)
            model.fit(X_train, y_train)

            # 在验证集上评估
            y_val_pred = model.predict(X_val)
            val_rmse = np.sqrt(mean_squared_error(y_val, y_val_pred))
            val_rmse_list.append(val_rmse)

        # 计算平均验证集RMSE
        avg_val_rmse = np.mean(val_rmse_list)

        # 更新最佳参数
        if avg_val_rmse < best_random_val_rmse:
            best_random_val_rmse = avg_val_rmse
            best_random_params = param_dict

    # 使用最佳参数训练最终模型
    best_random_model = XGBRegressor(**best_random_params)
    best_random_model.fit(X_train_val_random, y_train_val_random)

    # 在测试集上评估
    y_test_pred_random = best_random_model.predict(X_test_random)
    best_random_test_rmse = np.sqrt(mean_squared_error(y_test_random, y_test_pred_random))

    # 在训练集上评估
    y_train_pred_random = best_random_model.predict(X_train_val_random)
    best_random_train_rmse = np.sqrt(mean_squared_error(y_train_val_random, y_train_pred_random))

    # 确保结果文件夹存在
    os.makedirs(output_folder_path, exist_ok=True)

    # 读取测试数据并预测
    test_files = [f for f in os.listdir(test_folder_path) if f.endswith('.xlsx')]

    for file in test_files:
        try:
            test_data = pd.read_excel(os.path.join(test_folder_path, file))

            # 检查测试数据中的缺失值
            print(f"\n测试文件 {file} 缺失值统计:")
            missing_counts = test_data[features].isnull().sum()
            print(missing_counts[missing_counts > 0])

            # 创建测试数据的缺失值指示特征
            for col in features:
                if col in test_data.columns:
                    test_data[f'{col}_missing'] = test_data[col].isnull().astype(int)
                else:
                    test_data[f'{col}_missing'] = 1  # 特征不存在时标记为缺失

            # 确保测试特征与训练一致
            X_test = test_data[features]

            # 使用最佳模型进行预测
            predicted_rwt_space = best_space_model.predict(X_test)
            predicted_rwt_time = best_time_model.predict(X_test)
            predicted_rwt_random = best_random_model.predict(X_test)

            # 保存预测结果
            test_data['Predicted_RWT_Space'] = predicted_rwt_space
            test_data['Predicted_RWT_Time'] = predicted_rwt_time
            test_data['Predicted_RWT_Random'] = predicted_rwt_random
            output_path = os.path.join(output_folder_path, file)
            test_data.to_excel(output_path, index=False)
            print(f"已处理并保存: {output_path}")

        except Exception as e:
            print(f"处理文件 {file} 时出错: {str(e)}")

    # 输出所有交叉验证结果
    print("\n" + "=" * 50)
    print("三种交叉验证方式下的三数据集RMSE结果")
    print("=" * 50)

    print("\n【空间交叉验证】")
    print(f"训练集RMSE: {best_space_train_rmse:.4f}")
    print(f"验证集RMSE: {best_val_rmse:.4f}")
    print(f"测试集RMSE: {best_space_test_rmse:.4f}")
    print(f"最佳参数: {best_space_params}")

    print("\n【时间交叉验证】")
    print(f"训练集RMSE: {best_time_train_rmse:.4f}")
    print(f"验证集RMSE: {best_time_val_rmse:.4f}")
    print(f"测试集RMSE: {best_time_test_rmse:.4f}")
    print(f"最佳参数: {best_time_params}")

    print("\n【随机交叉验证】")
    print(f"训练集RMSE: {best_random_train_rmse:.4f}")
    print(f"验证集RMSE: {best_random_val_rmse:.4f}")
    print(f"测试集RMSE: {best_random_test_rmse:.4f}")
    print(f"最佳参数: {best_random_params}")

    # 计算特征贡献度
    calculate_feature_contributions(
        X, best_space_model, best_time_model, best_random_model, output_folder_path
    )


def calculate_feature_contributions(X, space_model, time_model, random_model, output_folder):
    print("\n正在计算特征贡献度...")
    try:
        # 初始化SHAP解释器
        space_explainer = shap.TreeExplainer(space_model)
        time_explainer = shap.TreeExplainer(time_model)
        random_explainer = shap.TreeExplainer(random_model)

        # 计算SHAP值（使用前1000样本加速计算）
        sample_idx = np.random.choice(len(X), min(1000, len(X)), replace=False)
        X_sample = X.iloc[sample_idx] if isinstance(X, pd.DataFrame) else X[sample_idx]

        space_shap_values = space_explainer.shap_values(X_sample)
        time_shap_values = time_explainer.shap_values(X_sample)
        random_shap_values = random_explainer.shap_values(X_sample)

        # 整理特征贡献度结果
        feature_contributions = pd.DataFrame({
            'Feature': X.columns,
            'Space_Mean_Abs_SHAP': np.abs(space_shap_values).mean(axis=0),
            'Space_Rank': np.argsort(np.argsort(-np.abs(space_shap_values).mean(axis=0))) + 1,
            'Space_Direction': np.sign(space_shap_values).mean(axis=0),
            'Time_Mean_Abs_SHAP': np.abs(time_shap_values).mean(axis=0),
            'Time_Rank': np.argsort(np.argsort(-np.abs(time_shap_values).mean(axis=0))) + 1,
            'Time_Direction': np.sign(time_shap_values).mean(axis=0),
            'Random_Mean_Abs_SHAP': np.abs(random_shap_values).mean(axis=0),
            'Random_Rank': np.argsort(np.argsort(-np.abs(random_shap_values).mean(axis=0))) + 1,
            'Random_Direction': np.sign(random_shap_values).mean(axis=0),
        })

        # 计算综合贡献度
        feature_contributions['Mean_Contribution'] = (
                                                             feature_contributions['Space_Mean_Abs_SHAP'] +
                                                             feature_contributions['Time_Mean_Abs_SHAP'] +
                                                             feature_contributions['Random_Mean_Abs_SHAP']
                                                     ) / 3
        feature_contributions['Overall_Rank'] = (
                np.argsort(np.argsort(-feature_contributions['Mean_Contribution'])) + 1
        )
        total_contribution = feature_contributions['Mean_Contribution'].sum()
        feature_contributions['Contribution_Percentage'] = (
                feature_contributions['Mean_Contribution'] / total_contribution * 100
        )

        # 保存特征贡献度结果
        output_path = os.path.join(output_folder, 'Feature_Contributions.xlsx')
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            feature_contributions.to_excel(writer, sheet_name='Contributions', index=False)
            workbook = writer.book
            sheet = workbook['Contributions']
            # 美化表格
            header_font = Font(bold=True)
            for cell in sheet[1]:
                cell.font = header_font
            for col in sheet.columns:
                max_len = max(len(str(cell.value)) for cell in col) + 2
                sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max_len, 30)

        print(f"特征贡献度已保存到: {output_path}")
        generate_contribution_plots(feature_contributions, output_folder)

    except Exception as e:
        print(f"计算特征贡献度时出错: {str(e)}")


def generate_contribution_plots(contributions, output_folder):
    plot_folder = os.path.join(output_folder, 'Contribution_Plots')
    os.makedirs(plot_folder, exist_ok=True)
    top_n = min(20, len(contributions))
    top_features = contributions.sort_values('Mean_Contribution', ascending=False).head(top_n)

    # 总体贡献度图
    plt.figure(figsize=(10, 8))
    colors = np.where(top_features['Space_Direction'] > 0, '#3498db', '#e74c3c')
    plt.barh(top_features['Feature'], top_features['Mean_Contribution'], color=colors)
    plt.xlabel('平均SHAP值绝对值（贡献度）')
    plt.ylabel('特征')
    plt.title('特征总体贡献度排名（Top 20）')
    plt.grid(axis='x', linestyle='--', alpha=0.7)
    plt.tight_layout()
    plt.savefig(os.path.join(plot_folder, 'Overall_Contribution.png'), dpi=300)
    plt.close()

    # 分模型贡献度图
    for model_type in ['Space', 'Time', 'Random']:
        plt.figure(figsize=(10, 8))
        colors = np.where(top_features[f'{model_type}_Direction'] > 0, '#3498db', '#e74c3c')
        plt.barh(
            top_features['Feature'],
            top_features[f'{model_type}_Mean_Abs_SHAP'],
            color=colors
        )
        plt.xlabel('平均SHAP值绝对值（贡献度）')
        plt.ylabel('特征')
        plt.title(f'{model_type}模型特征贡献度排名（Top 20）')
        plt.grid(axis='x', linestyle='--', alpha=0.7)
        plt.tight_layout()
        plt.savefig(os.path.join(plot_folder, f'{model_type}_Contribution.png'), dpi=600)
        plt.close()


if __name__ == "__main__":
    # 路径配置
    train_file_path = "E:\\huai_river\\Station_RWT\\Extracted_matching_with_temp.xlsx"
    test_folder_path = 'E:\\huai_river\\Huairiver_GEE_data\\Daily_data\\old\\none' #11-test-1
    output_folder_path = "E:\huai_river\Huairiver_GEE_data\Daily_data\old\\test2"  # result-0805_3
    train_and_predict(train_file_path, test_folder_path, output_folder_path)

    # 路径配置
    # train_file_path = "E:\\huai_river\\Station_RWT\\Extracted_matching_with_temp.xlsx"
    # test_folder_path = 'E:\\huai_river\\Huairiver_GEE_data\\Daily_data\\old\\All_site_LST\\1-train_data'
    # output_folder_path = "E:\\huai_river\\Huairiver_GEE_data\\Daily_data\\old\\All_site_LST\\2-result"  # result-0805_3
    # train_and_predict(train_file_path, test_folder_path, output_folder_path)

    # 路径配置
    # train_file_path = "E:\\huai_river\\Station_RWT\\Extracted_matching_with_temp_LST.xlsx"
    # test_folder_path = 'E:\\Global_river_temp\\Space_RWT\\3_binary_LST\\zone1'
    # output_folder_path = "E:\\Global_river_temp\\Space_RWT\\4_result\\zone1"
    # train_and_predict(train_file_path, test_folder_path, output_folder_path)


    # train_file_path = "E:\\huai_river\\Station_RWT\\Extracted_matching_with_temp.xlsx"
    # test_folder_path = 'E:\\huai_river\\Accuracy\\0608_overhaul\\comment1\\5_multi_rank_v2\\rank5' #11-test-1
    # output_folder_path = "E:\\huai_river\\Accuracy\\0608_overhaul\\comment1\\5_multi_rank_v2\\rank5_pred"  # result-0805_3
    # train_and_predict(train_file_path, test_folder_path, output_folder_path)








# ==============================   气温最大 最小值  ===============================================
import os
import pandas as pd
from datetime import datetime


def calculate_at_mean_average(folder_path, output_file):
    """
    计算文件夹中所有日期命名的Excel文件中"AT_mean"列的平均值，并汇总输出

    参数:
        folder_path: 包含日期命名Excel文件的文件夹路径
        output_file: 输出结果的Excel文件路径
    """
    # 存储结果的列表
    results = []

    # 遍历文件夹中的所有文件
    for filename in os.listdir(folder_path):
        # 筛选出符合日期格式的Excel文件（如"2019-01-01.xlsx"）
        if filename.endswith('.xlsx') and len(filename) == 15:
            try:
                # 提取文件名中的日期
                date_str = filename[:10]
                date = datetime.strptime(date_str, '%Y-%m-%d').date()

                # 读取Excel文件
                file_path = os.path.join(folder_path, filename)
                df = pd.read_excel(file_path)

                # 检查是否存在"AT_mean"列
                if "AT_mean" not in df.columns:
                    print(f"警告：文件 {filename} 中未找到'AT_mean'列，已跳过")
                    continue

                # 计算平均值（忽略空值）
                at_mean_avg = df["AT_mean"].mean(skipna=True)

                # 添加到结果列表
                results.append({
                    "日期": date,
                    "AT_mean平均值": at_mean_avg
                })

                print(f"已处理: {filename}，平均值: {at_mean_avg:.4f}")

            except Exception as e:
                print(f"处理文件 {filename} 时出错: {e}")
                continue

    if not results:
        print("错误：未找到有效数据或所有文件处理失败")
        return

    # 将结果转换为DataFrame并按日期排序
    result_df = pd.DataFrame(results)
    result_df = result_df.sort_values(by="日期").reset_index(drop=True)

    # 保存结果到Excel
    try:
        result_df.to_excel(output_file, index=False)
        print(f"\n处理完成！结果已保存至: {output_file}")
        print(f"共处理 {len(result_df)} 个有效文件")
    except Exception as e:
        print(f"保存结果失败: {e}")


# if __name__ == "__main__":
#     # 请根据实际情况修改以下路径
#     input_folder = r"E:\\huai_river\\Huairiver_GEE_data\\Daily_data\\old\\result-0718"  # 例如：r"C:\data\daily_files"
#     output_excel = r"E:\\huai_river\\Huairiver_GEE_data\\Daily_data\\old\\AT_average.xlsx"  # 例如：r"C:\data\AT_mean_averages.xlsx"
#
#     # 检查输入文件夹是否存在
#     if not os.path.isdir(input_folder):
#         print(f"错误：输入文件夹不存在 - {input_folder}")
#     else:
        # calculate_at_mean_average(input_folder, output_excel)